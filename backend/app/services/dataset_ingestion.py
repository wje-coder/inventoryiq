"""CSV/Excel parsing, validation, column normalization, and type inference.

Design note on where row data lives: this module writes a *normalized*
CSV file to disk (standardized column names, values as strings) rather
than inserting every row into Postgres as an individual record. With
uploads that can run to tens of thousands of rows, synchronous
per-row inserts inside an HTTP request would be slow enough to justify
background job infrastructure - which the spec explicitly says to avoid
unless necessary. Postgres instead holds what it's best suited for here:
dataset metadata, column schema, and validation findings. Preview and any
future analytics read rows back out of the normalized file via pandas.
"""

import csv
import dataclasses
from pathlib import Path
from zipfile import BadZipFile

import pandas as pd

from app.core.config import get_settings
from app.models.dataset import ColumnDataType, DatasetFileType, FindingSeverity

settings = get_settings()

_BOOL_TRUE = {"true", "1", "yes", "y"}
_BOOL_FALSE = {"false", "0", "no", "n"}
_BOOL_VALUES = _BOOL_TRUE | _BOOL_FALSE


class IngestionError(Exception):
    """A fatal, user-facing validation failure. `code` is machine-readable;
    `findings` are zero or more specific (severity, code, message, row,
    column) tuples suitable for persisting as DatasetValidationFinding rows.
    """

    def __init__(self, code: str, message: str, findings: list["Finding"] | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.findings = findings or []


@dataclasses.dataclass(frozen=True)
class Finding:
    severity: FindingSeverity
    code: str
    message: str
    row_number: int | None = None
    column_name: str | None = None


@dataclasses.dataclass
class ColumnResult:
    source_name: str
    normalized_name: str
    position: int
    inferred_type: ColumnDataType
    nullable: bool
    sample_values: list[str]


@dataclasses.dataclass
class IngestionResult:
    row_count: int
    column_count: int
    columns: list[ColumnResult]
    findings: list[Finding]
    normalized_relative_path: str


def _read_raw_csv_header(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        try:
            return next(reader)
        except StopIteration as exc:
            raise IngestionError("EMPTY_FILE", "The uploaded file has no content.") from exc


def _scan_csv_for_malformed_rows(path: Path, expected_fields: int) -> list[int]:
    """Return 1-indexed data-row numbers whose field count doesn't match
    the header. A "malformed CSV" is one where at least one non-blank
    row has a different number of fields than the header row.
    """
    malformed_rows: list[int] = []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        next(reader, None)  # header, already validated separately
        for row_number, row in enumerate(reader, start=1):
            if not row or all(cell.strip() == "" for cell in row):
                continue  # blank trailing lines are not "malformed"
            if len(row) != expected_fields:
                malformed_rows.append(row_number)
    return malformed_rows


def _read_raw_xlsx_header(path: Path) -> list[str]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, KeyError, OSError) as exc:
        raise IngestionError(
            "UNREADABLE_EXCEL", "The Excel file could not be read. It may be corrupted."
        ) from exc

    try:
        if not workbook.sheetnames:
            raise IngestionError(
                "UNSUPPORTED_SPREADSHEET_STRUCTURE", "The workbook contains no worksheets."
            )
        sheet = workbook[workbook.sheetnames[0]]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row is None or all(cell is None for cell in first_row):
            raise IngestionError(
                "UNSUPPORTED_SPREADSHEET_STRUCTURE",
                "The first worksheet has no header row.",
            )
        return [str(cell).strip() if cell is not None else "" for cell in first_row]
    finally:
        workbook.close()


def detect_duplicate_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    duplicates: list[str] = []
    for name in headers:
        seen[name] = seen.get(name, 0) + 1
        if seen[name] == 2:
            duplicates.append(name)
    return duplicates


def _raise_if_duplicate_headers(headers: list[str]) -> None:
    duplicates = sorted(set(detect_duplicate_headers(headers)))
    if not duplicates:
        return
    raise IngestionError(
        "DUPLICATE_COLUMN_NAMES",
        f"Duplicate column names found: {', '.join(duplicates)}.",
        findings=[
            Finding(
                FindingSeverity.ERROR,
                "DUPLICATE_COLUMN_NAME",
                f"Column '{name}' appears more than once.",
                column_name=name,
            )
            for name in duplicates
        ],
    )


def normalize_column_name(name: str, position: int) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in name.strip().lower())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    cleaned = cleaned.strip("_")
    return cleaned or f"column_{position}"


def _dedupe_normalized_names(names: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for name in names:
        counts[name] = counts.get(name, 0) + 1
        result.append(name if counts[name] == 1 else f"{name}_{counts[name]}")
    return result


def _infer_series_type(series: pd.Series) -> ColumnDataType:
    non_null = series.dropna().astype(str).str.strip()
    non_null = non_null[non_null != ""]
    if non_null.empty:
        return ColumnDataType.UNKNOWN

    numeric = pd.to_numeric(non_null, errors="coerce")
    if numeric.notna().all():
        if (numeric % 1 == 0).all():
            return ColumnDataType.INTEGER
        return ColumnDataType.FLOAT

    lowered = non_null.str.lower()
    if lowered.isin(_BOOL_VALUES).all():
        return ColumnDataType.BOOLEAN

    parsed_dates = pd.to_datetime(non_null, errors="coerce", format="mixed")
    if parsed_dates.notna().all():
        has_time_component = (
            (parsed_dates.dt.hour != 0)
            | (parsed_dates.dt.minute != 0)
            | (parsed_dates.dt.second != 0)
        ).any()
        return ColumnDataType.DATETIME if has_time_component else ColumnDataType.DATE

    return ColumnDataType.STRING


def _sample_values(series: pd.Series, limit: int) -> list[str]:
    non_null = series.dropna().astype(str).str.strip()
    non_null = non_null[non_null != ""]
    seen: list[str] = []
    for value in non_null:
        if value not in seen:
            seen.append(value)
        if len(seen) >= limit:
            break
    return seen


def _parse_dataframe(path: Path, file_type: DatasetFileType) -> pd.DataFrame:
    if file_type == DatasetFileType.CSV:
        try:
            return pd.read_csv(path, dtype=str, keep_default_na=True, encoding="utf-8-sig")
        except pd.errors.EmptyDataError as exc:
            raise IngestionError("EMPTY_FILE", "The uploaded file has no content.") from exc
        except (pd.errors.ParserError, UnicodeDecodeError, ValueError) as exc:
            raise IngestionError("MALFORMED_CSV", "The CSV file could not be parsed.") from exc

    try:
        return pd.read_excel(path, dtype=str, engine="openpyxl")
    except (BadZipFile, KeyError, ValueError, OSError) as exc:
        raise IngestionError(
            "UNREADABLE_EXCEL", "The Excel file could not be read. It may be corrupted."
        ) from exc


def ingest_dataset_file(
    path: Path, file_type: DatasetFileType, normalized_output_path: Path
) -> IngestionResult:
    """Validate and parse an already-saved dataset file, returning
    inferred column metadata and writing a normalized CSV alongside it.
    Raises IngestionError for any fatal validation failure.
    """
    findings: list[Finding] = []

    if file_type == DatasetFileType.CSV:
        raw_headers = _read_raw_csv_header(path)
        _raise_if_duplicate_headers(raw_headers)

        malformed_rows = _scan_csv_for_malformed_rows(path, expected_fields=len(raw_headers))
        if malformed_rows:
            raise IngestionError(
                "MALFORMED_CSV",
                f"{len(malformed_rows)} row(s) have a field count that doesn't match the header.",
                findings=[
                    Finding(
                        FindingSeverity.ERROR,
                        "MALFORMED_ROW",
                        "Row has an unexpected number of fields.",
                        row_number=row,
                    )
                    for row in malformed_rows[:20]
                ],
            )
    else:
        raw_headers = _read_raw_xlsx_header(path)
        _raise_if_duplicate_headers(raw_headers)

    dataframe = _parse_dataframe(path, file_type)
    dataframe = dataframe.dropna(axis=0, how="all")

    if dataframe.shape[1] == 0:
        raise IngestionError("NO_USABLE_ROWS", "The file has no columns.")
    if dataframe.shape[0] == 0:
        raise IngestionError("NO_USABLE_ROWS", "The file has no usable data rows.")

    source_names = [str(c).strip() for c in dataframe.columns]
    normalized_names = _dedupe_normalized_names(
        [normalize_column_name(name, i) for i, name in enumerate(source_names)]
    )

    columns: list[ColumnResult] = []
    paired_names = zip(source_names, normalized_names, strict=True)
    for position, (source_name, normalized_name) in enumerate(paired_names):
        series = dataframe.iloc[:, position]
        inferred_type = _infer_series_type(series)
        nullable = bool(series.isna().any() or (series.astype(str).str.strip() == "").any())
        columns.append(
            ColumnResult(
                source_name=source_name,
                normalized_name=normalized_name,
                position=position,
                inferred_type=inferred_type,
                nullable=nullable,
                sample_values=_sample_values(series, settings.dataset_column_sample_size),
            )
        )

    normalized_df = dataframe.copy()
    normalized_df.columns = normalized_names
    normalized_output_path.parent.mkdir(parents=True, exist_ok=True)
    normalized_df.to_csv(normalized_output_path, index=False)

    duplicate_data_rows = int(dataframe.duplicated().sum())
    if duplicate_data_rows:
        findings.append(
            Finding(
                FindingSeverity.WARNING,
                "DUPLICATE_ROWS",
                f"{duplicate_data_rows} duplicate row(s) detected and kept as-is.",
            )
        )

    return IngestionResult(
        row_count=int(dataframe.shape[0]),
        column_count=int(dataframe.shape[1]),
        columns=columns,
        findings=findings,
        normalized_relative_path=str(normalized_output_path),
    )
